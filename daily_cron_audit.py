import os
import io
import requests
import pandas as pd
import numpy as np
import openpyxl
from zoneinfo import ZoneInfo
from datetime import datetime, timedelta
import yfinance as yf

TIINGO_TOKEN = "bcffe3a5cf7eeef085e405cfa4a3e5691b976217"
EXCEL_FILE = "NQ_RADAR_Daily_Tracker.xlsx"

BENCHMARKS = ["QQQ", "NQ=F"]
MAG_7 = ["NVDA", "AAPL", "MSFT", "AMZN", "META", "GOOGL", "TSLA"]
STORAGE = ["MU", "WDC", "STX"]
WATCHLIST = ["QQQ"] + MAG_7 + STORAGE

def fetch_single_ticker_data(ticker):
    if ticker not in ["NQ=F"]:
        url = f"https://api.tiingo.com/tiingo/daily/{ticker.lower()}/prices"
        params = {
            "token": TIINGO_TOKEN,
            "startDate": (datetime.now() - timedelta(days=60)).strftime("%Y-%m-%d"),
            "columns": "date,open,high,low,close,volume"
        }
        try:
            r = requests.get(url, params=params, headers={'Content-Type': 'application/json'}, timeout=10)
            if r.status_code == 200 and len(r.json()) > 0:
                df = pd.DataFrame(r.json())
                df['Date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                df.set_index('Date', inplace=True)
                df.rename(columns={'open': 'Open', 'high': 'High', 'low': 'Low', 'close': 'Close', 'volume': 'Volume'}, inplace=True)
                return df[['Open', 'High', 'Low', 'Close', 'Volume']].dropna()
        except Exception:
            pass

    try:
        yf_df = yf.download(ticker, period="2mo", interval="1d", progress=False)
        if isinstance(yf_df.columns, pd.MultiIndex):
            yf_df.columns = yf_df.columns.get_level_values(0)
        return yf_df[['Open', 'High', 'Low', 'Close', 'Volume']].dropna()
    except Exception:
        return pd.DataFrame()

def append_to_excel(records):
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily_NQ_Radar_Log"
        headers = [
            "日期 (Date)", "标的 (Asset)", "收盘价 (Close)", "涨跌幅 (Chg %)", 
            "振幅 (Range $)", "相对量 (RVol)", "K线形态 (Candle Structure)", 
            "VPA量价信号 (VPA Signal)", "波动形态 (Crabel Model)", 
            "压缩状态 (Squeeze)", "一票否决 (Veto Status)", "预设操作 (Planned Action)", "复盘备注 (Notes)"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Daily_NQ_Radar_Log" not in wb.sheetnames:
        ws = wb.create_sheet(title="Daily_NQ_Radar_Log")
    else:
        ws = wb["Daily_NQ_Radar_Log"]

    existing_keys = set()
    for r in range(2, ws.max_row + 1):
        d_val = ws.cell(row=r, column=1).value
        a_val = ws.cell(row=r, column=2).value
        if d_val and a_val:
            existing_keys.add(f"{str(d_val)[:10]}_{str(a_val)}")

    row_idx = 2
    while ws.cell(row=row_idx, column=1).value is not None:
        row_idx += 1

    added = 0
    for rec in records:
        key = f"{rec['date']}_{rec['asset']}"
        if key in existing_keys:
            continue
        ws.cell(row=row_idx, column=1, value=rec['date'])
        ws.cell(row=row_idx, column=2, value=rec['asset'])
        ws.cell(row=row_idx, column=3, value=rec['close'])
        ws.cell(row=row_idx, column=4, value=rec['pct_chg'] / 100.0)
        ws.cell(row=row_idx, column=5, value=rec['range'])
        ws.cell(row=row_idx, column=6, value=rec['rvol'])
        ws.cell(row=row_idx, column=7, value=rec['candle'])
        ws.cell(row=row_idx, column=8, value=rec['vpa'])
        ws.cell(row=row_idx, column=9, value=rec['crabel'])
        ws.cell(row=row_idx, column=10, value=rec['squeeze'])
        ws.cell(row=row_idx, column=11, value=rec['veto'])
        ws.cell(row=row_idx, column=12, value=rec['action'])
        ws.cell(row=row_idx, column=13, value=rec['notes'])
        row_idx += 1
        added += 1

    wb.save(EXCEL_FILE)
    print(f"✅ 定格数据归档完成: 新增 {added} 条记录 (跳过重复)。")

def run_audit():
    print("🚀 启动每日 NQ RADAR 自动化审计任务...")
    records = []
    
    # 优先锚定 NQ 或 QQQ 最新交易日
    ref_df = fetch_single_ticker_data("QQQ")
    if ref_df.empty:
        print("🛑 无法获取基准数据，任务中止。")
        return
    
    target_idx = ref_df.index[-1]
    target_date_str = target_idx.strftime('%Y-%m-%d')
    print(f"📅 锁定定格目标交易日: {target_date_str}")

    for ticker in WATCHLIST:
        df = fetch_single_ticker_data(ticker)
        if df.empty or len(df) < 8:
            continue
            
        pos = df.index.get_loc(target_idx) if target_idx in df.index else len(df) - 1
        
        o = df['Open'].iloc[pos]
        c = df['Close'].iloc[pos]
        h = df['High'].iloc[pos]
        l = df['Low'].iloc[pos]
        prev_c = df['Close'].iloc[pos - 1] if pos > 0 else c
        prev_h = df['High'].iloc[pos - 1] if pos > 0 else h
        prev_l = df['Low'].iloc[pos - 1] if pos > 0 else l
        
        chg_pct = ((c - prev_c) / prev_c) * 100
        rng = h - l
        vol = df['Volume'].iloc[pos]
        vol_ma20 = df['Volume'].iloc[max(0, pos-20):pos].mean() if pos > 0 else vol
        rvol = vol / vol_ma20 if vol_ma20 > 0 else 1.0
        
        # Toby Crabel 模型
        past_ranges = (df['High'].iloc[max(0, pos-6):pos+1] - df['Low'].iloc[max(0, pos-6):pos+1]).values
        is_nr7 = len(past_ranges) == 7 and rng == np.min(past_ranges)
        is_inside = (h <= prev_h) and (l >= prev_l)
        
        if is_nr7:
            toby_status = "NR7 (7日极窄)"
            squeeze_status = "Tight Compression ⚡"
        elif is_inside:
            toby_status = "Inside Bar (孕线)"
            squeeze_status = "Tight Compression ⚡"
        else:
            toby_status = "Normal"
            squeeze_status = "Neutral"

        # VPA 核心判定
        lower_wick = min(o, c) - l
        upper_wick = h - max(o, c)
        candle_struct = "Normal Body"
        vpa_signal = "Normal Volume"
        veto_status = "PASS ✅"
        action_signal = "Watch & Wait 🔒"
        
        if lower_wick > (rng * 0.35) and rvol >= 1.2:
            vpa_signal = "STOPPING VOLUME 🛑"
            candle_struct = "Lower Wick Rejection"
            action_signal = "Long Trigger 🚀"
        elif upper_wick > (rng * 0.35) and rvol >= 1.2:
            vpa_signal = "TOPPING OUT 🛑"
            candle_struct = "Upper Wick Rejection"
            veto_status = "VETO 🛑"
            action_signal = "Sit on Hands (一票否决) 🔒"
        elif rvol >= 1.8 and abs(chg_pct) < 0.3:
            vpa_signal = "Volume Anomaly (量大滞涨)"
            veto_status = "VETO 🛑"
            action_signal = "Sit on Hands (一票否决) 🔒"

        records.append({
            "date": target_date_str,
            "asset": ticker,
            "close": round(c, 2),
            "pct_chg": round(chg_pct, 2),
            "range": round(rng, 2),
            "rvol": round(rvol, 2),
            "candle": candle_struct,
            "vpa": vpa_signal,
            "crabel": toby_status,
            "squeeze": squeeze_status,
            "veto": veto_status,
            "action": action_signal,
            "notes": f"RVol: {rvol:.2f}x | 自动定时审计"
        })

    append_to_excel(records)

if __name__ == "__main__":
    run_audit()