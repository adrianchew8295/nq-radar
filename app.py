import os
import io
import requests
import pandas as pd
import numpy as np
import openpyxl
from zoneinfo import ZoneInfo
from datetime import datetime, timedelta
import streamlit as st
import yfinance as yf

# ==============================================================================
# 🎨 1. 页面基础配置 (自适应布局)
# ==============================================================================
st.set_page_config(
    page_title="NQmain 跨资产时光机与量价雷达",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 专属授权 Token (Tiingo 机构级直连 - 已内置)
TIINGO_TOKEN = "bcffe3a5cf7eeef085e405cfa4a3e5691b976217"
EXCEL_FILE = "NQ_RADAR_Daily_Tracker.xlsx"

# 专属标的池 (已剔除退市的 SNDK，换入活跃存储龙头 STX)
BENCHMARKS = ["QQQ", "NQ=F"]
MAG_7 = ["NVDA", "AAPL", "MSFT", "AMZN", "META", "GOOGL", "TSLA"]
STORAGE = ["MU", "WDC", "STX"]
WATCHLIST = ["QQQ"] + MAG_7 + STORAGE

# ==============================================================================
# 🌐 2. 自动抓取 Invesco 官方 QQQ 每日成分股
# ==============================================================================
@st.cache_data(ttl=86400)
def fetch_invesco_qqq_holdings():
    url = "https://www.invesco.com/us/financial-products/etfs/holdings/main/holdings-0.csv?audienceType=Investor&action=download&ticker=QQQ"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        res = requests.get(url, headers=headers, timeout=8)
        if res.status_code == 200:
            df = pd.read_csv(io.StringIO(res.text))
            if 'Holding Ticker' in df.columns:
                tickers = df['Holding Ticker'].dropna().astype(str).str.strip().tolist()
                valid = [t for t in tickers if t.isalpha() and len(t) <= 5]
                return valid[:40]
    except Exception:
        pass
    return [
        "NVDA", "AAPL", "MSFT", "AMZN", "META", "GOOGL", "GOOG", "TSLA", "AVGO", "COST",
        "NFLX", "AMD", "QCOM", "ADBE", "LIN", "TXN", "PEP", "AMGN", "ISRG", "INTU",
        "CMCSA", "HON", "BKNG", "AMAT", "VRTX", "LRCX", "ADI", "PANW", "MU", "PLTR",
        "KLAC", "SNPS", "CDNS", "MDLZ", "CRWD", "MAR", "ORLY", "CTAS", "NXPI", "FTNT"
    ]

TOP_WEIGHTS = fetch_invesco_qqq_holdings()

# ==============================================================================
# 📡 3. 机构级数据抓取引擎 (Tiingo + yfinance 双重兜底)
# ==============================================================================
@st.cache_data(ttl=1800)
def fetch_single_ticker_data(ticker):
    if ticker not in ["NQ=F"]:
        url = f"https://api.tiingo.com/tiingo/daily/{ticker.lower()}/prices"
        params = {
            "token": TIINGO_TOKEN,
            "startDate": (datetime.now() - timedelta(days=730)).strftime("%Y-%m-%d"),
            "columns": "date,open,high,low,close,volume"
        }
        headers = {'Content-Type': 'application/json'}
        try:
            r = requests.get(url, params=params, headers=headers, timeout=8)
            if r.status_code == 200 and len(r.json()) > 0:
                data = r.json()
                df = pd.DataFrame(data)
                df['Date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                df.set_index('Date', inplace=True)
                df.rename(columns={
                    'open': 'Open', 'high': 'High', 'low': 'Low',
                    'close': 'Close', 'volume': 'Volume'
                }, inplace=True)
                return df[['Open', 'High', 'Low', 'Close', 'Volume']].dropna()
        except Exception:
            pass

    # 备用引擎
    try:
        yf_df = yf.download(ticker, period="2y", interval="1d", progress=False)
        if isinstance(yf_df.columns, pd.MultiIndex):
            yf_df.columns = yf_df.columns.get_level_values(0)
        return yf_df[['Open', 'High', 'Low', 'Close', 'Volume']].dropna()
    except Exception:
        return pd.DataFrame()

@st.cache_data(ttl=1800)
def load_all_market_data():
    all_data = {}
    combined_tickers = list(set(WATCHLIST + TOP_WEIGHTS + ["NQ=F"]))
    for t in combined_tickers:
        df = fetch_single_ticker_data(t)
        if len(df) > 0:
            all_data[t] = df
    return all_data

# ==============================================================================
# 🛑 4. 铁律门禁：原生标准美东时区验证 (夏冬令时自适应)
# ==============================================================================
def check_market_lockout(target_dt):
    now_et = datetime.now(ZoneInfo("America/New_York"))
    today_et = now_et.date()
    if target_dt == today_et:
        # 美东 16:15 之后视为完全定格
        if now_et.time() < datetime.strptime("16:15", "%H:%M").time():
            return False, now_et.strftime('%H:%M:%S')
    return True, ""

# ==============================================================================
# 📊 5. Excel 自动排重追加写入算法 (防丢防错)
# ==============================================================================
def append_to_excel_tracker(log_records):
    # 如果文件不存在，自动初始化新建
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily_NQ_Radar_Log"
        headers = [
            "日期 (Date)", "标的 (Asset)", "收盘价 (Close)", "涨跌幅 (Chg %)", 
            "振幅 (Range $)", "相对量 (RVol)", "K线形态 (Candle Structure)", 
            "VPA量价信号 (VPA Signal)", "波动形态 (Crabel Model)", 
            "压缩状态 (Squeeze)", "一票否决 (Veto Status)", "预设操作 (Planned Action)", "复盘备注 (Notes)"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Daily_NQ_Radar_Log" not in wb.sheetnames:
            ws = wb.create_sheet(title="Daily_NQ_Radar_Log")
        else:
            ws = wb["Daily_NQ_Radar_Log"]
        
        # 读取已存在的记录，防止重复写入
        existing_keys = set()
        for r in range(2, ws.max_row + 1):
            d_val = ws.cell(row=r, column=1).value
            a_val = ws.cell(row=r, column=2).value
            if d_val and a_val:
                existing_keys.add(f"{str(d_val)[:10]}_{str(a_val)}")
                
        # 寻找首个空行
        row_idx = 2
        while ws.cell(row=row_idx, column=1).value is not None:
            row_idx += 1
            
        added_count = 0
        for rec in log_records:
            key = f"{rec['date']}_{rec['asset']}"
            if key in existing_keys:
                continue
                
            ws.cell(row=row_idx, column=1, value=rec['date'])
            ws.cell(row=row_idx, column=2, value=rec['asset'])
            ws.cell(row=row_idx, column=3, value=rec['close'])
            ws.cell(row=row_idx, column=4, value=rec['pct_chg'] / 100.0)
            ws.cell(row=row_idx, column=5, value=rec['range'])
            ws.cell(row=row_idx, column=6, value=rec['rvol'])
            ws.cell(row=row_idx, column=7, value=rec['candle'])
            ws.cell(row=row_idx, column=8, value=rec['vpa'])
            ws.cell(row=row_idx, column=9, value=rec['crabel'])
            ws.cell(row=row_idx, column=10, value=rec['squeeze'])
            ws.cell(row=row_idx, column=11, value=rec['veto'])
            ws.cell(row=row_idx, column=12, value=rec['action'])
            ws.cell(row=row_idx, column=13, value=rec['notes'])
            
            row_idx += 1
            added_count += 1
            
        wb.save(EXCEL_FILE)
        return True, f"成功归档 {added_count} 条定格数据到 {EXCEL_FILE}！(已自动跳过重复记录)"
    except Exception as e:
        return False, f"写入 Excel 失败: {str(e)}"

# ==============================================================================
# 🎛️ 6. 侧边栏控制面板
# ==============================================================================
st.sidebar.header("🕹️ 时光机控制面板")
query_mode = st.sidebar.radio(
    "选择查询维度:",
    ["📅 指定历史日期", "🎯 NQmain 目标点位反查", "🏆 NQmain 历史最高点 (ATH)"]
)

target_price_input = None
target_date_input = None

if query_mode == "📅 指定历史日期":
    target_date_input = st.sidebar.date_input("选择查询交易日", datetime.now().date())
elif query_mode == "🎯 NQmain 目标点位反查":
    target_price_input = st.sidebar.number_input("输入 NQmain 点位", min_value=10000, max_value=50000, value=22000, step=100)

st.sidebar.markdown("---")
st.sidebar.success("⚡ 数据引擎: **Tiingo + VPA 算法已激活 ✅**")
st.sidebar.caption("💡 **操盘风控底线**: 坚持只认定格 DAILY CLOSE，严禁盘中虚假数据！")

# ==============================================================================
# 🧠 7. 核心算法与呈现引擎
# ==============================================================================
try:
    all_market_data = load_all_market_data()
    ref_df = all_market_data.get('NQ=F', all_market_data.get('QQQ'))
    
    if ref_df is None or len(ref_df) == 0:
        st.error("⚠️ 无法连接行情引擎，请检查网络或刷新重试。")
        st.stop()

    if query_mode == "🏆 NQmain 历史最高点 (ATH)":
        target_idx = ref_df['High'].idxmax()
    elif query_mode == "🎯 NQmain 目标点位反查":
        matched = ref_df[ref_df['High'] >= target_price_input]
        target_idx = matched.index[-1] if len(matched) > 0 else (ref_df['Close'] - target_price_input).abs().idxmin()
    else:
        t_dt = pd.to_datetime(target_date_input)
        valid_dates = ref_df.index[ref_df.index <= t_dt]
        target_idx = valid_dates[-1] if len(valid_dates) > 0 else ref_df.index[-1]

    target_date_str = target_idx.strftime('%Y-%m-%d')
    target_date_obj = target_idx.date()

    # 门禁检查
    is_closed, current_et_time = check_market_lockout(target_date_obj)
    if not is_closed:
        st.error(f"""
        🛑 **门禁拦截报错 (LOCKOUT ERROR): 拒绝处理未定格数据！**
        * **目标日期**: `{target_date_str}`
        * **当前美东时间**: `{current_et_time} ET` (未到 16:15 收盘定格)
        * **核心铁律**: DAILY CLOSE 尚未定格，严禁盘中复盘！请于美股收盘后再试。
        """)
        st.stop()

    # NQ 与 QQQ 宏观水位
    nq_df = all_market_data.get('NQ=F', ref_df)
    nq_close = nq_df['Close'].loc[target_idx] if target_idx in nq_df.index else nq_df['Close'].iloc[-1]
    loc_pos_nq = nq_df.index.get_loc(target_idx) if target_idx in nq_df.index else len(nq_df)-1
    nq_prev = nq_df['Close'].iloc[loc_pos_nq - 1] if loc_pos_nq > 0 else nq_close
    nq_chg_pct = ((nq_close - nq_prev) / nq_prev) * 100
    nq_ath = nq_df['High'].loc[:target_idx].max() if target_idx in nq_df.index else nq_df['High'].max()
    nq_drawdown = ((nq_close - nq_ath) / nq_ath) * 100

    qqq_df = all_market_data.get('QQQ')
    qqq_close = qqq_df['Close'].loc[target_idx] if target_idx in qqq_df.index else qqq_df['Close'].iloc[-1]
    loc_pos_q = qqq_df.index.get_loc(target_idx) if target_idx in qqq_df.index else len(qqq_df)-1
    qqq_prev = qqq_df['Close'].iloc[loc_pos_q - 1] if loc_pos_q > 0 else qqq_close
    qqq_chg_pct = ((qqq_close - qqq_prev) / qqq_prev) * 100

    # 页面标题
    st.title("📡 NQmain 跨资产量价雷达与时光机")
    st.caption(f"📅 锁定定格交易日: **{target_date_str}** | 门禁审计: **DAILY CLOSE 已定格 ✅** | 引擎: **Tiingo + VPA 算法 ⚡**")

    col1, col2, col3 = st.columns(3)
    col1.metric("🎯 NQmain 期货主力", f"{nq_close:,.2f}", f"{nq_chg_pct:+.2f}%")
    col2.metric("📈 QQQ 纳指基准", f"${qqq_close:.2f}", f"{qqq_chg_pct:+.2f}%")
    col3.metric("🏛️ 大盘历史位阶", f"距 ATH {nq_drawdown:+.2f}%", "极高位冲顶" if nq_drawdown > -3 else "正常洗盘区")

    st.markdown("---")

    # 维度一：今日谁在拉盘 TOP 5
    st.subheader("🔥 维度一：拉盘动量归因 TOP 5 (Volume-Weighted Momentum)")
    contrib_list = []
    for t in TOP_WEIGHTS:
        if t not in all_market_data: continue
        df_t = all_market_data[t]
        if target_idx not in df_t.index: continue
        pos = df_t.index.get_loc(target_idx)
        if pos == 0: continue
        
        c = df_t['Close'].iloc[pos]
        prev_c = df_t['Close'].iloc[pos - 1]
        vol = df_t['Volume'].iloc[pos]
        dollar_vol = c * vol
        chg_pct = ((c - prev_c) / prev_c) * 100
        contrib_pts = (chg_pct / 100.0) * (dollar_vol / 1e9) * 2.5
        
        contrib_list.append({
            "代码": t,
            "当日收盘价": f"${c:.2f}",
            "当日涨跌%": chg_pct,
            "成交额($B)": dollar_vol / 1e9,
            "拉盘动量得分": contrib_pts
        })

    df_contrib = pd.DataFrame(contrib_list).sort_values(by="拉盘动量得分", ascending=False).head(5)
    df_contrib_show = df_contrib.copy()
    df_contrib_show['当日涨跌%'] = df_contrib_show['当日涨跌%'].apply(lambda x: f"{x:+.2f}%")
    df_contrib_show['成交额($B)'] = df_contrib_show['成交额($B)'].apply(lambda x: f"${x:.2f}B")
    df_contrib_show['拉盘动量得分'] = df_contrib_show['拉盘动量得分'].apply(lambda x: f"{x:+.1f} 🚀")
    st.dataframe(df_contrib_show.reset_index(drop=True), use_container_width=True)

    # 维度二：专属关注池 (VPA + Toby Crabel 核心算法)
    st.subheader("🎯 维度二：专属关注池 (VPA 量价 + Toby 压缩 + 一票否决审计)")
    watch_rows = []
    excel_records = []
    copy_lines = []

    for ticker in WATCHLIST:
        if ticker not in all_market_data:
            continue
            
        df_t = all_market_data[ticker]
        pos = df_t.index.get_loc(target_idx) if target_idx in df_t.index else len(df_t) - 1
        
        o = df_t['Open'].iloc[pos]
        c = df_t['Close'].iloc[pos]
        h = df_t['High'].iloc[pos]
        l = df_t['Low'].iloc[pos]
        prev_c = df_t['Close'].iloc[pos - 1] if pos > 0 else c
        prev_h = df_t['High'].iloc[pos - 1] if pos > 0 else h
        prev_l = df_t['Low'].iloc[pos - 1] if pos > 0 else l
        
        chg_pct = ((c - prev_c) / prev_c) * 100
        rng = h - l
        
        vol = df_t['Volume'].iloc[pos]
        vol_ma20 = df_t['Volume'].iloc[max(0, pos-20):pos].mean() if pos > 0 else vol
        rvol = vol / vol_ma20 if vol_ma20 > 0 else 1.0
        
        ath_p = df_t['High'].loc[:df_t.index[pos]].max()
        drawdown = ((c - ath_p) / ath_p) * 100
        
        # 1. Toby Crabel 波动形态算法
        past_ranges = (df_t['High'].iloc[max(0, pos-6):pos+1] - df_t['Low'].iloc[max(0, pos-6):pos+1]).values
        is_nr7 = len(past_ranges) == 7 and rng == np.min(past_ranges)
        is_nr4 = len(past_ranges) >= 4 and rng == np.min(past_ranges[-4:])
        is_inside = (h <= prev_h) and (l >= prev_l)
        avg_5_rng = np.mean(past_ranges[-5:]) if len(past_ranges) >= 5 else rng
        is_expansion = rng > (1.3 * avg_5_rng) and (c > prev_h or c < prev_l)
        
        if is_nr7:
            toby_status = "NR7 (7日极窄)"
            squeeze_status = "Tight Compression ⚡"
        elif is_inside:
            toby_status = "Inside Bar (孕线)"
            squeeze_status = "Tight Compression ⚡"
        elif is_nr4:
            toby_status = "NR4 (4日收缩)"
            squeeze_status = "Tight Compression ⚡"
        elif is_expansion:
            toby_status = "Expansion (剧烈扩张)"
            squeeze_status = "Expanded"
        else:
            toby_status = "Normal"
            squeeze_status = "Neutral"

        # 2. VPA 量价核心算法 (Effort vs Result)
        lower_wick = min(o, c) - l
        upper_wick = h - max(o, c)
        candle_struct = "Normal Body"
        vpa_signal = "Normal Volume"
        veto_status = "PASS ✅"
        action_signal = "Watch & Wait 🔒"
        
        if lower_wick > (rng * 0.35) and rvol >= 1.2:
            vpa_signal = "STOPPING VOLUME 🛑"
            candle_struct = "Lower Wick Rejection"
            action_signal = "Long Trigger 🚀"
        elif upper_wick > (rng * 0.35) and rvol >= 1.2:
            vpa_signal = "TOPPING OUT 🛑"
            candle_struct = "Upper Wick Rejection"
            veto_status = "VETO 🛑"
            action_signal = "Sit on Hands (一票否决) 🔒"
        elif rvol >= 1.8 and abs(chg_pct) < 0.3:
            vpa_signal = "Volume Anomaly (量大滞涨)"
            veto_status = "VETO 🛑"
            action_signal = "Sit on Hands (一票否决) 🔒"
        elif is_expansion and chg_pct > 1.5:
            action_signal = "Long Trigger 🚀"

        cat = "基准" if ticker in BENCHMARKS else ("存储" if ticker in STORAGE else "7巨头")
        
        watch_rows.append({
            "代码": ticker,
            "板块": cat,
            "当日收盘": f"${c:.2f}" if ticker != "NQ=F" else f"{c:,.2f}",
            "涨跌%": chg_pct,
            "量比(20MA)": f"{rvol:.2f}x",
            "距自身ATH%": drawdown,
            "VPA 量价状态": vpa_signal,
            "TOBY 形态": toby_status,
            "一票否决审查": veto_status,
            "操盘指令": action_signal
        })
        
        excel_records.append({
            "date": target_date_str,
            "asset": ticker,
            "close": round(c, 2),
            "pct_chg": round(chg_pct, 2),
            "range": round(rng, 2),
            "rvol": round(rvol, 2),
            "candle": candle_struct,
            "vpa": vpa_signal,
            "crabel": toby_status,
            "squeeze": squeeze_status,
            "veto": veto_status,
            "action": action_signal,
            "notes": f"RVol: {rvol:.2f}x | ATH回撤: {drawdown:.1f}% | 算法自动生成"
        })
        
        copy_lines.append(f" • {ticker:5s}: 收盘 {c:.2f} ({chg_pct:+.2f}%) | VPA: {vpa_signal} | TOBY: {toby_status} | 否决: {veto_status} -> {action_signal}")

    df_watch = pd.DataFrame(watch_rows)
    st.dataframe(df_watch.reset_index(drop=True), use_container_width=True)

    # 维度三：Excel 历史归档按钮 (1-CLICK ARCHIVE 📥)
    st.markdown("---")
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        if st.button("📥 一键追加归档至 Excel 历史账本", type="primary", use_container_width=True):
            success, msg = append_to_excel_tracker(excel_records)
            if success:
                st.success(f"✅ {msg}")
            else:
                st.error(f"🛑 {msg}")
                
    with col_info:
        st.caption(f"💡 点击将自动将 `{target_date_str}` 的全部审计数据写入 `{EXCEL_FILE}`，并自动进行防重复过滤。")

    # 维度四：一键复制给 Gemini 审计报告
    st.markdown("---")
    st.subheader("📋 一键极速复制 (粘贴给 Gemini 执行 431 穿透)")
    
    top1 = df_contrib.iloc[0]['代码'] if len(df_contrib) > 0 else "N/A"
    top1_chg = df_contrib.iloc[0]['当日涨跌%'] if len(df_contrib) > 0 else 0
    
    copy_text_block = f"""================================================================================
📡 NQ RADAR 量价与波动率定格报告 (DAILY CLOSE 定格版)
📅 查询基准日: {target_date_str} | 🎯 NQmain: {nq_close:,.2f} ({nq_chg_pct:+.2f}%) | 📈 QQQ: ${qqq_close:.2f} ({qqq_chg_pct:+.2f}%)
--------------------------------------------------------------------------------
【🔥 今日拉盘动量 TOP 1】: {top1} ({top1_chg:+.2f}%)

【🎯 专属关注池 VPA 与 TOBY 风控状态】
""" + "\n".join(copy_lines) + """
================================================================================
"""
    st.code(copy_text_block, language="text")

except Exception as e:
    st.error(f"⚠️ 运行出现异常: {str(e)}")
